# =====================================================
# 모듈 4 - 주석 취합 + 교차 검증 프로그램
# 여러 팀에서 받은 주석 파일을 하나로 합치고,
# 팀별 수치가 기준 결산명세서와 일치하는지 자동으로 검사합니다.
# =====================================================

import pandas as pd                          # 표 데이터를 다루는 도구
from openpyxl.styles import PatternFill, Font, Alignment  # 엑셀 스타일 도구
import os                                     # 파일/폴더 경로 처리 도구

# -------------------------------------------------------
# [설정] 여기서 원하는 값을 바꾸세요!
# -------------------------------------------------------

# 각 팀에서 받은 주석 파일 경로 목록
# 파일은 반드시 "계정코드", "계정명", "금액" 열을 포함해야 합니다
FOOTNOTE_FILES = [
    # "data/input/주석_영업팀_2026년1분기.xlsx",
    # "data/input/주석_관리팀_2026년1분기.xlsx",
    # "data/input/주석_재무팀_2026년1분기.xlsx",
]

# 기준이 되는 분기 결산명세서 경로 (module2 결과물)
MASTER_FILE = "data/output/분기결산명세서_2026년1분기.xlsx"

# 결과 저장 경로
OUTPUT_MERGED     = "data/output/주석_통합_2026년1분기.xlsx"
OUTPUT_VALIDATION = "data/output/교차검증_결과_2026년1분기.xlsx"

# 주석 파일의 열 이름 (실제 파일에 맞게 수정하세요)
COL_ACCOUNT = "계정코드"  # 계정코드 열
COL_NAME    = "계정명"   # 계정명 열
COL_AMOUNT  = "금액"     # 금액 열

# 교차 검증 허용 오차 (반올림 차이 무시)
TOLERANCE = 1  # 1원 이하 차이는 정상으로 처리

# -------------------------------------------------------
# [색상 설정]
# -------------------------------------------------------
COLOR_RED    = "FF4444"  # 불일치 항목 (빨간색)
COLOR_GREEN  = "C6EFCE"  # 일치 항목 (연두색)
COLOR_YELLOW = "FFD700"  # 주의 항목 (노란색)
COLOR_HEADER = "4472C4"  # 헤더 (파란색)


# -------------------------------------------------------
# [함수] 팀별 주석 파일을 모두 불러옵니다
# -------------------------------------------------------
def load_all_footnotes(file_list):
    """
    여러 팀의 주석 파일을 딕셔너리(팀명: 데이터) 형태로 불러옵니다.
    입력: 파일 경로 목록
    출력: {팀이름: DataFrame} 형태의 딕셔너리
    """
    print("  📂 팀별 주석 파일 불러오는 중...")

    if not file_list:
        print("  [주의] FOOTNOTE_FILES 설정에 파일 경로가 없습니다.")
        print("         설정 부분에 각 팀의 주석 파일 경로를 추가하세요.")
        return {}

    footnotes = {}  # 팀별 데이터를 담을 빈 딕셔너리

    for file_path in file_list:
        if not os.path.exists(file_path):
            print(f"    [주의] 파일 없음 → 건너뜀: {file_path}")
            continue

        # 파일 이름에서 팀 이름 추출 (예: "주석_영업팀_2026년1분기.xlsx" → "영업팀")
        filename = os.path.basename(file_path)
        team_name = filename.replace("주석_", "").split("_")[0]  # 팀 이름 추출

        # 파일 읽기
        try:
            df = pd.read_excel(file_path)

            # 필요한 열이 있는지 확인
            for col in [COL_ACCOUNT, COL_AMOUNT]:
                if col not in df.columns:
                    print(f"    [오류] '{team_name}' 파일에 '{col}' 열이 없습니다.")
                    continue

            df[COL_ACCOUNT] = df[COL_ACCOUNT].astype(str)  # 계정코드 문자열로 통일
            footnotes[team_name] = df
            print(f"    [완료] {team_name}: {len(df)}개 항목 불러옴")

        except Exception as e:
            print(f"    [오류] {team_name} 파일 읽기 실패: {e}")

    return footnotes


# -------------------------------------------------------
# [함수] 팀별 데이터를 하나의 표로 합칩니다
# -------------------------------------------------------
def merge_footnotes(footnotes_dict):
    """
    팀별로 분리된 데이터를 계정코드 기준으로 하나의 표로 통합합니다.
    입력: {팀이름: DataFrame} 딕셔너리
    출력: 통합된 DataFrame
    """
    if not footnotes_dict:
        return pd.DataFrame()

    print("  🔗 주석 데이터 통합 중...")

    # 첫 번째 팀 데이터를 기준으로 시작
    team_names = list(footnotes_dict.keys())
    result = footnotes_dict[team_names[0]][[COL_ACCOUNT]].copy()

    # 팀명이 있으면 계정명도 가져오기
    if COL_NAME in footnotes_dict[team_names[0]].columns:
        result[COL_NAME] = footnotes_dict[team_names[0]][COL_NAME]

    # 각 팀의 금액 열을 추가 (열 이름: "팀명_금액")
    for team, df in footnotes_dict.items():
        team_col = f"{team}_금액"
        if COL_AMOUNT in df.columns:
            team_data = df[[COL_ACCOUNT, COL_AMOUNT]].rename(columns={COL_AMOUNT: team_col})
            result = result.merge(team_data, on=COL_ACCOUNT, how="outer")

    result = result.fillna(0)  # 없는 항목은 0으로 채우기
    print(f"  [완료] {len(result)}개 항목 통합 완료")
    return result


# -------------------------------------------------------
# [함수] 기준 결산명세서와 교차 검증합니다
# -------------------------------------------------------
def cross_validate(merged_df, master_file):
    """
    통합된 주석 수치를 기준 결산명세서와 비교합니다.
    입력: 통합 DataFrame, 기준 파일 경로
    출력: 검증 결과 DataFrame
    """
    print("  🔍 기준 수치와 교차 검증 중...")

    if not os.path.exists(master_file):
        print(f"  [오류] 기준 파일을 찾을 수 없습니다: {master_file}")
        return merged_df

    # 기준 파일 읽기 (결산명세서 시트)
    master_df = pd.read_excel(master_file, sheet_name="결산명세서")
    master_df[COL_ACCOUNT] = master_df[COL_ACCOUNT].astype(str)
    master_df = master_df[[COL_ACCOUNT, "당기금액"]].rename(columns={"당기금액": "기준금액"})

    # 기준 수치와 합치기
    result = merged_df.merge(master_df, on=COL_ACCOUNT, how="outer")
    result = result.fillna(0)

    # 팀별 금액 열 찾기
    team_cols = [c for c in result.columns if c.endswith("_금액")]

    # 각 팀별로 기준과의 차이 계산
    discrepancy_cols = []
    for team_col in team_cols:
        diff_col = team_col.replace("_금액", "_차이")
        result[diff_col] = result[team_col] - result["기준금액"]
        discrepancy_cols.append(diff_col)

    # 전체 불일치 여부 판단 (허용 오차 초과 시 불일치)
    def check_match(row):
        for diff_col in discrepancy_cols:
            if abs(row.get(diff_col, 0)) > TOLERANCE:
                return "❌ 불일치"
        return "✅ 일치"

    result["검증결과"] = result.apply(check_match, axis=1)

    match_count = (result["검증결과"] == "✅ 일치").sum()
    mismatch_count = (result["검증결과"] == "❌ 불일치").sum()
    print(f"  [완료] 일치 {match_count}건 / 불일치 {mismatch_count}건")

    return result


# -------------------------------------------------------
# [함수] 불일치 항목을 추출합니다
# -------------------------------------------------------
def find_discrepancies(validation_df):
    """
    검증 결과에서 허용 오차를 초과하는 불일치 항목만 추출합니다.
    입력: 검증 결과 DataFrame
    출력: 불일치 항목 DataFrame
    """
    if "검증결과" not in validation_df.columns:
        return pd.DataFrame()

    discrepancies = validation_df[validation_df["검증결과"] == "❌ 불일치"].copy()
    return discrepancies


# -------------------------------------------------------
# [함수] 통합 파일을 저장합니다
# -------------------------------------------------------
def save_merged_output(merged_df, output_path):
    """
    통합된 주석 데이터를 엑셀 파일로 저장합니다.
    입력: 통합 DataFrame, 저장 경로
    출력: 없음 (파일 저장)
    """
    if merged_df.empty:
        print("  [주의] 통합할 데이터가 없습니다.")
        return

    print(f"  💾 통합 파일 저장 중... ({output_path})")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged_df.to_excel(writer, sheet_name="통합주석", index=False)
        ws = writer.sheets["통합주석"]

        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    print(f"  [완료] '{output_path}' 저장 완료!")


# -------------------------------------------------------
# [함수] 검증 결과 보고서를 저장합니다
# -------------------------------------------------------
def save_validation_report(validation_df, discrepancies, output_path):
    """
    교차 검증 결과를 시트 3개짜리 엑셀 파일로 저장합니다.
    입력: 전체 검증 DataFrame, 불일치 DataFrame, 저장 경로
    출력: 없음 (파일 저장)
    """
    print(f"  💾 검증 결과 보고서 저장 중... ({output_path})")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    total = len(validation_df) if not validation_df.empty else 0
    mismatch = len(discrepancies) if not discrepancies.empty else 0
    match = total - mismatch
    passed = "✅ 통과" if mismatch == 0 else "❌ 불일치 있음"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # === 시트1: 검증 요약 ===
        summary_data = {
            "항목": ["총 검증 항목 수", "일치 항목 수", "불일치 항목 수", "최종 검증 결과"],
            "결과": [total, match, mismatch, passed],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="검증요약", index=False)
        ws1 = writer.sheets["검증요약"]

        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        # 결과 행에 색상 (통과=초록, 실패=빨강)
        for row in ws1.iter_rows(min_row=2):
            val = str(row[1].value)
            if "❌" in val:
                fill = PatternFill("solid", fgColor=COLOR_RED)
            elif "✅" in val:
                fill = PatternFill("solid", fgColor=COLOR_GREEN)
            else:
                continue
            for cell in row:
                cell.fill = fill

        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = max_len + 4

        # === 시트2: 불일치 목록 ===
        if not discrepancies.empty:
            discrepancies.to_excel(writer, sheet_name="불일치목록", index=False)
            ws2 = writer.sheets["불일치목록"]

            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)

            red_fill = PatternFill("solid", fgColor=COLOR_RED)
            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = red_fill

            note_row = len(discrepancies) + 3
            ws2.cell(row=note_row, column=1, value="※ 빨간색 행 = 기준 수치와 불일치 → 담당자 확인 필요")
            ws2.cell(row=note_row, column=1).font = Font(color="FF0000", italic=True)

            for col in ws2.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws2.column_dimensions[col[0].column_letter].width = max_len + 4
        else:
            pd.DataFrame([{"내용": "불일치 항목 없음 ✅"}]).to_excel(
                writer, sheet_name="불일치목록", index=False)

        # === 시트3: 전체 검증 결과 ===
        if not validation_df.empty:
            validation_df.to_excel(writer, sheet_name="전체검증", index=False)
            ws3 = writer.sheets["전체검증"]

            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)

            # 결과 열에 따라 색상 (일치=초록, 불일치=빨강)
            result_col = None
            for i, cell in enumerate(ws3[1]):
                if cell.value == "검증결과":
                    result_col = i
                    break

            if result_col is not None:
                green_fill = PatternFill("solid", fgColor=COLOR_GREEN)
                red_fill = PatternFill("solid", fgColor=COLOR_RED)
                for row in ws3.iter_rows(min_row=2):
                    val = str(row[result_col].value)
                    if "✅" in val:
                        for cell in row:
                            cell.fill = green_fill
                    elif "❌" in val:
                        for cell in row:
                            cell.fill = red_fill

            for col in ws3.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws3.column_dimensions[col[0].column_letter].width = max_len + 4

    print(f"  [완료] '{output_path}' 저장 완료!")


# -------------------------------------------------------
# [메인] 프로그램 실행 시 이 부분이 동작합니다
# -------------------------------------------------------
def main():
    print("=" * 55)
    print("  주석 취합 + 교차 검증 프로그램 시작!")
    print("=" * 55)
    print(f"  기준 파일: {MASTER_FILE}")
    print(f"  취합 파일 수: {len(FOOTNOTE_FILES)}개")
    print()

    # 1단계: 팀별 주석 파일 불러오기
    footnotes = load_all_footnotes(FOOTNOTE_FILES)
    print()

    # 2단계: 하나의 표로 통합
    merged_df = merge_footnotes(footnotes)
    print()

    # 3단계: 기준 수치와 교차 검증
    validation_df = cross_validate(merged_df, MASTER_FILE)
    print()

    # 4단계: 불일치 항목 추출
    discrepancies = find_discrepancies(validation_df)
    print()

    # 5단계: 통합 파일 저장
    save_merged_output(merged_df, OUTPUT_MERGED)

    # 6단계: 검증 결과 보고서 저장
    save_validation_report(validation_df, discrepancies, OUTPUT_VALIDATION)

    # 최종 요약
    print()
    print("=" * 55)
    if not discrepancies.empty:
        print(f"  ⚠️  불일치 {len(discrepancies)}건 발견 → '{OUTPUT_VALIDATION}' 확인 필요")
    else:
        print("  ✅ 모든 항목 검증 통과!")
    print(f"  통합 파일: {OUTPUT_MERGED}")
    print(f"  검증 결과: {OUTPUT_VALIDATION}")
    print("=" * 55)


# 프로그램 시작점
if __name__ == "__main__":
    main()
