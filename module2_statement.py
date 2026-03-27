# =====================================================
# 모듈 2 - 분기 결산명세서 자동화 프로그램
# 월 마감 보고서 3개를 합쳐서 분기 결산명세서를 만듭니다.
# 전기(작년 같은 분기)와 비교하여 증감액/증감률도 계산합니다.
# =====================================================

import pandas as pd                          # 표 데이터를 다루는 도구
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # 엑셀 스타일 도구
import os                                     # 파일/폴더 경로 처리 도구

# -------------------------------------------------------
# [설정] 여기서 원하는 값을 바꾸세요!
# -------------------------------------------------------

# 몇 분기인지, 어느 연도인지
QUARTER      = 1      # 분기 (1~4)
TARGET_YEAR  = 2026   # 당기 연도
PRIOR_YEAR   = 2025   # 전기 연도 (비교 대상)

# 당기 월 마감 보고서 파일 경로 3개 (module1_closing.py 결과물)
CURRENT_FILES = [
    "data/output/월마감_보고서_2026년01월.xlsx",
    "data/output/월마감_보고서_2026년02월.xlsx",
    "data/output/월마감_보고서_2026년03월.xlsx",
]

# 전기 분기 결산명세서 파일 경로 (비교용, 없으면 None으로 두세요)
PRIOR_FILE = None  # 예: "data/output/분기결산명세서_2025년1분기.xlsx"

# 결과 저장 경로
OUTPUT_FILE = f"data/output/분기결산명세서_{TARGET_YEAR}년{QUARTER}분기.xlsx"

# 변동률 이상값 기준: 이 비율 이상 변동하면 노란색 표시
VARIANCE_THRESHOLD = 0.20  # 0.20 = 20%

# -------------------------------------------------------
# [색상 설정]
# -------------------------------------------------------
COLOR_YELLOW = "FFD700"  # 이상 항목 (노란색)
COLOR_HEADER = "4472C4"  # 헤더 (파란색)
COLOR_SUBHEADER = "8EA9C1"  # 서브 헤더 (연파란)
COLOR_TOTAL  = "D9E1F2"  # 합계 행 (연보라)


# -------------------------------------------------------
# [함수] 월 마감 보고서 파일들을 불러옵니다
# -------------------------------------------------------
def load_monthly_reports(file_list):
    """
    module1에서 만든 월 마감 보고서 엑셀 파일들을 불러와 하나로 합칩니다.
    입력: 파일 경로 목록 (리스트)
    출력: 통합된 pandas DataFrame
    """
    print("  📂 월 마감 보고서 불러오는 중...")
    all_data = []  # 각 월 데이터를 담을 빈 목록

    for file_path in file_list:
        if not os.path.exists(file_path):
            print(f"    [주의] 파일 없음 → 건너뜀: {file_path}")
            continue

        # 시트1(집계표)에서 계정코드, 계정명, 당월금액만 읽기
        df = pd.read_excel(file_path, sheet_name="집계표")

        # 파일 이름에서 월 추출 (예: "월마감_보고서_2026년02월.xlsx" → "02월")
        month_label = os.path.basename(file_path).replace("월마감_보고서_", "").replace(".xlsx", "")

        # 열 이름에 월 정보 추가 (예: "당월금액" → "2월금액")
        month_num = month_label.replace(f"{TARGET_YEAR}년", "").replace("월", "")
        df = df.rename(columns={"당월금액": f"{month_num}월"})

        all_data.append(df)
        print(f"    [완료] {month_label} 불러옴 ({len(df)}개 계정)")

    if not all_data:
        print("  [오류] 불러올 수 있는 파일이 없습니다.")
        return None

    # 여러 월 데이터를 계정코드 기준으로 합치기
    result = all_data[0][["계정코드", "계정명"]].copy()  # 계정코드, 계정명은 첫 파일에서
    for df in all_data:
        # 금액 열만 추출해서 붙이기
        amount_cols = [c for c in df.columns if c not in ["계정코드", "계정명"]]
        result = result.merge(df[["계정코드"] + amount_cols], on="계정코드", how="outer")

    result = result.fillna(0)  # 없는 데이터는 0으로 채우기
    print(f"  [완료] 총 {len(result)}개 계정 통합 완료")
    return result


# -------------------------------------------------------
# [함수] 분기 합산합니다
# -------------------------------------------------------
def aggregate_quarterly(df):
    """
    3개월 금액을 합산하여 분기 합계를 계산합니다.
    입력: 월별 금액이 포함된 DataFrame
    출력: 분기 합계가 추가된 DataFrame
    """
    print("  ➕ 분기 합산 중...")

    # 금액 열만 찾기 (열 이름이 "X월" 형태인 것)
    amount_cols = [c for c in df.columns if c.endswith("월") and c != "계정명"]

    # 분기 합계 계산
    df["당기금액"] = df[amount_cols].sum(axis=1)  # 가로로 더하기

    print(f"  [완료] 분기 합계: {df['당기금액'].sum():,.0f}원")
    return df


# -------------------------------------------------------
# [함수] 전기와 비교합니다
# -------------------------------------------------------
def compare_with_prior(current_df, prior_file):
    """
    당기 분기와 전기 분기를 비교하여 증감액/증감률을 계산합니다.
    입력: 당기 DataFrame, 전기 파일 경로 (없으면 None)
    출력: 비교표 DataFrame
    """
    print("  📊 전기 대비 비교 중...")

    result = current_df[["계정코드", "계정명", "당기금액"]].copy()

    if prior_file and os.path.exists(prior_file):
        # 전기 파일에서 당기금액(→전기금액) 읽기
        prior_df = pd.read_excel(prior_file, sheet_name="결산명세서")
        if "당기금액" in prior_df.columns:
            prior_df = prior_df[["계정코드", "당기금액"]].rename(columns={"당기금액": "전기금액"})
            result = result.merge(prior_df, on="계정코드", how="left")
            result["전기금액"] = result["전기금액"].fillna(0)
        else:
            result["전기금액"] = 0
            print("    [주의] 전기 파일에서 금액 열을 찾지 못했습니다.")
    else:
        result["전기금액"] = 0
        if prior_file:
            print(f"    [주의] 전기 파일 없음 → 전기금액 0으로 처리: {prior_file}")
        else:
            print("    [주의] 전기 파일 미설정 → 전기금액 0으로 처리")

    # 증감액, 증감률 계산
    result["증감액"] = result["당기금액"] - result["전기금액"]
    result["증감률"] = result.apply(
        lambda row: (row["증감액"] / row["전기금액"]) if row["전기금액"] != 0 else None,
        axis=1
    )

    print("  [완료] 전기 비교 완료")
    return result


# -------------------------------------------------------
# [함수] 엑셀 파일로 저장합니다
# -------------------------------------------------------
def save_statement(comparison_df, monthly_df, output_path):
    """
    결산명세서를 시트 3개짜리 엑셀 파일로 저장합니다.
    입력: 비교표 DataFrame, 월별 내역 DataFrame, 저장 경로
    출력: 없음 (파일 저장)
    """
    print(f"  💾 엑셀 파일 저장 중... ({output_path})")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # === 시트1: 결산명세서 ===
        comparison_df.to_excel(writer, sheet_name="결산명세서", index=False)
        ws1 = writer.sheets["결산명세서"]

        # 헤더 스타일
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 증감률 이상 항목 노란색 표시
        yellow_fill = PatternFill("solid", fgColor=COLOR_YELLOW)
        variance_col = None
        for i, cell in enumerate(ws1[1]):
            if cell.value == "증감률":
                variance_col = i + 1  # openpyxl은 1부터 시작
                break

        for row in ws1.iter_rows(min_row=2):
            if variance_col:
                variance_cell = row[variance_col - 1]
                val = variance_cell.value
                if val is not None and abs(val) > VARIANCE_THRESHOLD:
                    for cell in row:
                        cell.fill = yellow_fill

        # 숫자 서식 (금액: 콤마, 증감률: 퍼센트)
        for row in ws1.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ["C", "D", "E"]:  # 당기, 전기, 증감액
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                elif cell.column_letter == "F":  # 증감률
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0%"

        # 이상 항목 안내 문구
        last_row = len(comparison_df) + 3
        ws1.cell(row=last_row, column=1,
                 value=f"※ 노란색 행: 전기 대비 {VARIANCE_THRESHOLD:.0%} 초과 변동 → 주석 작성 시 변동 사유 기재 필요")
        ws1.cell(row=last_row, column=1).font = Font(color="FF0000", italic=True)

        # 열 너비 자동 조정
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = max_len + 4

        # === 시트2: 월별 내역 ===
        monthly_df.to_excel(writer, sheet_name="월별내역", index=False)
        ws2 = writer.sheets["월별내역"]

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = max_len + 4

        # === 시트3: 검토사항 ===
        # 증감률 20% 초과 항목만 추출
        if "증감률" in comparison_df.columns:
            review_items = comparison_df[
                comparison_df["증감률"].notna() &
                (comparison_df["증감률"].abs() > VARIANCE_THRESHOLD)
            ].copy()
        else:
            review_items = pd.DataFrame()

        if len(review_items) > 0:
            review_items["검토사항"] = "[담당자 확인 필요] 변동 사유를 주석에 기재하세요"
            review_items.to_excel(writer, sheet_name="검토사항", index=False)
        else:
            pd.DataFrame([{"내용": "전기 대비 20% 초과 변동 항목 없음"}]).to_excel(
                writer, sheet_name="검토사항", index=False)

        ws3 = writer.sheets["검토사항"]
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col in ws3.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws3.column_dimensions[col[0].column_letter].width = max_len + 4

    print(f"  [완료] '{output_path}' 저장 완료!")


# -------------------------------------------------------
# [메인] 프로그램 실행 시 이 부분이 동작합니다
# -------------------------------------------------------
def main():
    print("=" * 55)
    print("  분기 결산명세서 자동화 프로그램 시작!")
    print("=" * 55)
    print(f"  대상: {TARGET_YEAR}년 {QUARTER}분기")
    print(f"  비교: {PRIOR_YEAR}년 {QUARTER}분기 (전기)")
    print()

    # 1단계: 월 마감 보고서 3개 불러오기
    monthly_df = load_monthly_reports(CURRENT_FILES)
    if monthly_df is None:
        return
    print()

    # 2단계: 분기 합산
    monthly_df = aggregate_quarterly(monthly_df)
    print()

    # 3단계: 전기 비교
    comparison_df = compare_with_prior(monthly_df, PRIOR_FILE)
    print()

    # 4단계: 엑셀 저장
    save_statement(comparison_df, monthly_df, OUTPUT_FILE)

    # 최종 요약
    total = comparison_df["당기금액"].sum()
    review_count = comparison_df["증감률"].notna() & (comparison_df["증감률"].abs() > VARIANCE_THRESHOLD)
    print()
    print("=" * 55)
    print(f"  {QUARTER}분기 합계: {total:,.0f}원")
    if review_count.sum() > 0:
        print(f"  ⚠️  검토 필요 항목 {review_count.sum()}건 → 시트3(검토사항) 확인 필요")
    print("=" * 55)


# 프로그램 시작점
if __name__ == "__main__":
    main()
