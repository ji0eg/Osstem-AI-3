# =====================================================
# 모듈 1 - 월 마감 자동화 프로그램
# ERP에서 내보낸 매출 데이터를 검증하고 계정별로 집계합니다.
# =====================================================

import pandas as pd                          # 표 데이터를 다루는 도구
from openpyxl.styles import PatternFill, Font, Alignment  # 엑셀 색상, 글씨, 정렬 스타일
from openpyxl.utils import get_column_letter  # 열 번호를 A, B, C... 로 변환
import os                                     # 파일/폴더 경로 처리 도구

# -------------------------------------------------------
# [설정] 여기서 원하는 값을 바꾸세요!
# -------------------------------------------------------

# ERP에서 내보낸 원시 데이터 파일 경로
INPUT_FILE = "data/input/erp_매출_2026년03월.csv"

# 결과를 저장할 엑셀 파일 경로
OUTPUT_FILE = "data/output/월마감_보고서_2026년03월.xlsx"

# 대상 연도와 월
TARGET_YEAR  = 2026
TARGET_MONTH = 3

# ERP 파일의 열 이름 (실제 ERP 파일에 맞게 수정하세요)
COL_DATE    = "거래일자"    # 날짜 열 이름
COL_ACCOUNT = "계정코드"   # 계정과목 코드 열 이름
COL_NAME    = "계정명"     # 계정과목 이름 열 이름
COL_AMOUNT  = "금액"       # 금액 열 이름
COL_NOTE    = "적요"       # 메모/설명 열 이름 (없으면 빈값 허용)

# 유효한 계정과목 코드 목록 (여기 없는 코드가 나오면 경고)
VALID_ACCOUNT_CODES = ["4010", "4020", "4030", "4110", "4120"]

# 이상값 기준: 전월 대비 이 비율 이상 변동하면 노란색으로 표시
ANOMALY_THRESHOLD = 0.30  # 0.30 = 30%

# 전월 마감 파일 경로 (비교용, 없으면 None으로 두세요)
PREV_MONTH_FILE = None  # 예: "data/output/월마감_보고서_2026년02월.xlsx"

# -------------------------------------------------------
# [색상 설정]
# -------------------------------------------------------
COLOR_YELLOW = "FFD700"  # 이상 항목 (노란색)
COLOR_RED    = "FF4444"  # 오류 항목 (빨간색)
COLOR_GREEN  = "C6EFCE"  # 정상 항목 (연두색)
COLOR_HEADER = "4472C4"  # 헤더 배경 (파란색)


# -------------------------------------------------------
# [함수] ERP 파일을 불러옵니다
# -------------------------------------------------------
def load_erp_data(file_path):
    """
    ERP에서 내보낸 CSV 또는 엑셀 파일을 pandas 표(DataFrame)로 불러옵니다.
    입력: 파일 경로 (문자열)
    출력: pandas DataFrame (표)
    """
    print(f"  📂 파일 불러오는 중... ({file_path})")

    # 파일이 존재하는지 확인
    if not os.path.exists(file_path):
        print(f"  [오류] 파일을 찾을 수 없습니다: {file_path}")
        print(f"         data/input/ 폴더에 ERP 파일을 넣어주세요.")
        return None

    # 파일 확장자에 따라 읽는 방법 선택
    if file_path.endswith(".csv"):
        df = pd.read_csv(file_path, encoding="utf-8-sig")  # CSV 파일 읽기
    elif file_path.endswith((".xlsx", ".xls")):
        df = pd.read_excel(file_path)  # 엑셀 파일 읽기
    else:
        print(f"  [오류] .csv 또는 .xlsx 파일만 지원합니다.")
        return None

    print(f"  [완료] {len(df)}개 행 불러오기 완료")
    return df


# -------------------------------------------------------
# [함수] 데이터 정합성을 검증합니다
# -------------------------------------------------------
def validate_data(df):
    """
    불러온 데이터에서 누락값, 중복, 잘못된 계정코드를 찾아냅니다.
    입력: pandas DataFrame
    출력: (검증된 DataFrame, 오류 목록 리스트)
    """
    print("  🔍 데이터 검증 중...")
    errors = []  # 오류를 모아둘 빈 목록

    # --- 필수 열이 있는지 확인 ---
    required_cols = [COL_DATE, COL_ACCOUNT, COL_AMOUNT]
    for col in required_cols:
        if col not in df.columns:
            print(f"  [오류] '{col}' 열이 파일에 없습니다. 설정의 열 이름을 확인하세요.")
            return df, [{"유형": "구조 오류", "내용": f"'{col}' 열 없음", "조치": "설정 수정 필요"}]

    # --- 금액이 비어있는 행 확인 ---
    missing_amount = df[df[COL_AMOUNT].isna()]  # 금액이 비어있는 행 추출
    if len(missing_amount) > 0:
        for idx in missing_amount.index:
            errors.append({
                "유형": "누락값",
                "행번호": idx + 2,  # 엑셀은 1부터 시작 + 헤더 행
                "내용": f"금액이 비어있습니다",
                "계정코드": df.loc[idx, COL_ACCOUNT] if COL_ACCOUNT in df.columns else "-",
                "조치": "[담당자 확인 필요]"
            })

    # --- 계정코드가 비어있는 행 확인 ---
    missing_account = df[df[COL_ACCOUNT].isna()]
    if len(missing_account) > 0:
        for idx in missing_account.index:
            errors.append({
                "유형": "누락값",
                "행번호": idx + 2,
                "내용": f"계정코드가 비어있습니다",
                "계정코드": "-",
                "조치": "[담당자 확인 필요]"
            })

    # --- 유효하지 않은 계정코드 확인 ---
    df[COL_ACCOUNT] = df[COL_ACCOUNT].astype(str)  # 코드를 문자열로 변환
    invalid_accounts = df[~df[COL_ACCOUNT].isin(VALID_ACCOUNT_CODES) & df[COL_ACCOUNT].notna()]
    if len(invalid_accounts) > 0:
        for idx in invalid_accounts.index:
            code = df.loc[idx, COL_ACCOUNT]
            errors.append({
                "유형": "코드 오류",
                "행번호": idx + 2,
                "내용": f"등록되지 않은 계정코드: {code}",
                "계정코드": code,
                "조치": "[담당자 확인 필요]"
            })

    # --- 중복 행 확인 (날짜 + 계정코드 + 금액이 완전히 동일한 경우) ---
    dup_mask = df.duplicated(subset=[COL_DATE, COL_ACCOUNT, COL_AMOUNT], keep=False)
    duplicates = df[dup_mask]
    if len(duplicates) > 0:
        errors.append({
            "유형": "중복 의심",
            "행번호": "-",
            "내용": f"날짜+계정코드+금액이 동일한 행 {len(duplicates)}개 발견",
            "계정코드": "-",
            "조치": "[담당자 확인 필요]"
        })

    # 오류 요약 출력
    if len(errors) == 0:
        print(f"  [완료] 검증 통과! 오류 없음")
    else:
        print(f"  [주의] 검증 결과: {len(errors)}개 오류 발견 → 시트2(검증결과) 확인 필요")

    return df, errors


# -------------------------------------------------------
# [함수] 전월 대비 이상 항목을 표시합니다
# -------------------------------------------------------
def flag_anomalies(current_df, prev_file):
    """
    전월 마감 파일과 비교하여 변동률이 기준을 초과하는 계정을 표시합니다.
    입력: 당월 DataFrame, 전월 파일 경로 (없으면 None)
    출력: 이상 계정코드 집합(set)
    """
    if prev_file is None or not os.path.exists(prev_file):
        print("  [주의] 전월 파일 없음 → 이상값 비교 생략")
        return set()  # 빈 집합 반환

    print(f"  📊 전월 대비 이상값 확인 중...")

    # 전월 집계 파일 불러오기 (시트1: 집계표)
    prev_agg = pd.read_excel(prev_file, sheet_name="집계표", index_col="계정코드")

    # 당월 집계
    current_agg = current_df.groupby(COL_ACCOUNT)[COL_AMOUNT].sum()

    anomaly_codes = set()  # 이상 항목 계정코드 모음

    for code in current_agg.index:
        current_val = current_agg[code]
        if code in prev_agg.index:
            prev_val = prev_agg.loc[code, "당월금액"]
            if prev_val != 0:
                change_rate = abs((current_val - prev_val) / prev_val)  # 변동률 계산
                if change_rate > ANOMALY_THRESHOLD:
                    anomaly_codes.add(code)
                    print(f"    [주의] 계정 {code}: 전월 대비 {change_rate:.1%} 변동")

    return anomaly_codes


# -------------------------------------------------------
# [함수] 계정과목별로 금액을 합산합니다
# -------------------------------------------------------
def aggregate_by_account(df):
    """
    계정코드별로 금액을 합산하여 요약 표를 만듭니다.
    입력: pandas DataFrame (원시 데이터)
    출력: 집계된 pandas DataFrame
    """
    print("  📋 계정별 집계 중...")

    # 금액 열을 숫자로 변환 (문자로 들어온 경우 대비)
    df[COL_AMOUNT] = pd.to_numeric(df[COL_AMOUNT], errors="coerce").fillna(0)

    # 계정코드 + 계정명으로 그룹화하여 금액 합산
    if COL_NAME in df.columns:
        agg = df.groupby([COL_ACCOUNT, COL_NAME])[COL_AMOUNT].sum().reset_index()
        agg.columns = ["계정코드", "계정명", "당월금액"]
    else:
        agg = df.groupby([COL_ACCOUNT])[COL_AMOUNT].sum().reset_index()
        agg.columns = ["계정코드", "당월금액"]
        agg["계정명"] = ""  # 계정명 열이 없으면 빈칸으로

    # 금액 기준 내림차순 정렬
    agg = agg.sort_values("당월금액", ascending=False).reset_index(drop=True)

    print(f"  [완료] {len(agg)}개 계정 집계 완료")
    return agg


# -------------------------------------------------------
# [함수] 엑셀 파일로 저장합니다
# -------------------------------------------------------
def save_closing_report(agg_df, errors, anomaly_codes, output_path):
    """
    집계 결과와 검증 오류를 엑셀 파일(시트 2개)로 저장합니다.
    입력: 집계 DataFrame, 오류 목록, 이상 계정코드 집합, 저장 경로
    출력: 없음 (파일 저장)
    """
    print(f"  💾 엑셀 파일 저장 중... ({output_path})")

    # 저장 폴더가 없으면 자동으로 만들기
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # === 시트1: 집계표 ===
        agg_df.to_excel(writer, sheet_name="집계표", index=False)
        ws1 = writer.sheets["집계표"]

        # 헤더 스타일 (파란 배경 + 흰 글씨)
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws1[1]:  # 1행 = 헤더
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 이상 항목 노란색 표시
        yellow_fill = PatternFill("solid", fgColor=COLOR_YELLOW)
        for row in ws1.iter_rows(min_row=2):  # 2행부터 데이터
            account_code = str(row[0].value)  # 첫 번째 열 = 계정코드
            if account_code in anomaly_codes:
                for cell in row:
                    cell.fill = yellow_fill

        # 금액 열 숫자 서식 (콤마 구분)
        for row in ws1.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

        # 이상 항목 안내 문구 추가 (집계표 아래)
        if anomaly_codes:
            last_row = len(agg_df) + 3
            ws1.cell(row=last_row, column=1, value="※ 노란색 행: 전월 대비 30% 초과 변동 → 담당자 확인 필요")
            ws1.cell(row=last_row, column=1).font = Font(color="FF0000", italic=True)

        # 열 너비 자동 조정
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = max_len + 4

        # === 시트2: 검증결과 ===
        if errors:
            errors_df = pd.DataFrame(errors)
        else:
            errors_df = pd.DataFrame([{"유형": "없음", "내용": "모든 항목 검증 통과 ✅", "조치": "-"}])

        errors_df.to_excel(writer, sheet_name="검증결과", index=False)
        ws2 = writer.sheets["검증결과"]

        # 헤더 스타일
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 오류 행 빨간색 표시
        red_fill = PatternFill("solid", fgColor=COLOR_RED)
        for row in ws2.iter_rows(min_row=2):
            error_type = str(row[0].value)
            if error_type != "없음":
                for cell in row:
                    cell.fill = red_fill

        # 열 너비 자동 조정
        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = max_len + 4

    print(f"  [완료] '{output_path}' 저장 완료!")


# -------------------------------------------------------
# [메인] 프로그램 실행 시 이 부분이 동작합니다
# -------------------------------------------------------
def main():
    print("=" * 55)
    print("  월 마감 자동화 프로그램 시작!")
    print("=" * 55)
    print(f"  대상: {TARGET_YEAR}년 {TARGET_MONTH}월")
    print(f"  입력 파일: {INPUT_FILE}")
    print()

    # 1단계: ERP 파일 불러오기
    df = load_erp_data(INPUT_FILE)
    if df is None:
        return  # 파일 없으면 종료

    print()

    # 2단계: 데이터 검증
    df, errors = validate_data(df)
    print()

    # 3단계: 전월 대비 이상값 확인
    anomaly_codes = flag_anomalies(df, PREV_MONTH_FILE)
    print()

    # 4단계: 계정별 집계
    agg_df = aggregate_by_account(df)
    print()

    # 5단계: 엑셀 저장
    save_closing_report(agg_df, errors, anomaly_codes, OUTPUT_FILE)

    # 최종 결과 요약
    print()
    print("=" * 55)
    print(f"  총 {len(agg_df)}개 계정, {df[COL_AMOUNT].sum():,.0f}원 집계 완료")
    if errors:
        print(f"  ⚠️  검증 오류 {len(errors)}건 → '{OUTPUT_FILE}' 시트2 확인 필요")
    if anomaly_codes:
        print(f"  ⚠️  이상 항목 {len(anomaly_codes)}건 → 노란색 행 확인 필요")
    print("=" * 55)


# 프로그램 시작점
if __name__ == "__main__":
    main()
